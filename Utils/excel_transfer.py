import os
import openpyxl
import shutil
import csv


def transfer(src_dataset_name, tar_excel_name, idx):
    if src_dataset_name == "THU": sub_num = 64
    elif src_dataset_name == "CAS": sub_num = 14
    elif src_dataset_name == "GIST": sub_num = 55

    method_name = ["rLDA", "HDCA", "xDAWNRG", "XGBDIM", "DeepConvNet", "EEGNet", "EEGInception", "PLNet", "PPNN", "DRL", "RP", "TS", "CPC", "MTCN"]
    ID = [i for i in range(1, sub_num+1)]
    result_excel = openpyxl.load_workbook(os.path.join(os.getcwd(), 'StatisResult', f'{src_dataset_name}', f'Result_{src_dataset_name}.xlsx'))
    sheet = result_excel[f'{tar_excel_name}']

    # # 生成第一列（编号）
    # for i in range(sub_num*len(method_name)):
    #     sheet['A'+ str(i+2)].value = i + 1
    # # 生成第二列（ID）
    # for i in range(len(method_name)):
    #     for j in range(sub_num):
    #         sheet['B'+ str(i*sub_num+j+2)] = j + 1
    # # 生成第三列（算法）
    # for i in range(len(method_name)):
    #     for j in range(sub_num):
    #         sheet['C'+ str(i*sub_num+j+2)] = method_name[i]
    # 从其他表格中复制出第四列（性能）
    for i in range(len(method_name)):
        tmp_sheet = result_excel[f'{method_name[i]}']
        for j in range(sub_num):
            if i == 0: sheet['B'+ str(j+2)] = tmp_sheet[idx + str(j+2)].value
            if i == 1: sheet['C'+ str(j+2)] = tmp_sheet[idx + str(j+2)].value
            if i == 2: sheet['D'+ str(j+2)] = tmp_sheet[idx + str(j+2)].value
            if i == 3: sheet['E'+ str(j+2)] = tmp_sheet[idx + str(j+2)].value
            if i == 4: sheet['F'+ str(j+2)] = tmp_sheet[idx + str(j+2)].value
            if i == 5: sheet['G'+ str(j+2)] = tmp_sheet[idx + str(j+2)].value
            if i == 6: sheet['H'+ str(j+2)] = tmp_sheet[idx + str(j+2)].value
            if i == 7: sheet['I'+ str(j+2)] = tmp_sheet[idx + str(j+2)].value
            if i == 8: sheet['J'+ str(j+2)] = tmp_sheet[idx + str(j+2)].value
            if i == 9: sheet['K'+ str(j+2)] = tmp_sheet[idx + str(j+2)].value
            if i == 10: sheet['L'+ str(j+2)] = tmp_sheet[idx + str(j+2)].value
            if i == 11: sheet['M'+ str(j+2)] = tmp_sheet[idx + str(j+2)].value
            if i == 12: sheet['N'+ str(j+2)] = tmp_sheet[idx + str(j+2)].value
            if i == 13: sheet['O'+ str(j+2)] = tmp_sheet[idx + str(j+2)].value

    # 保存文件
    result_excel.save(os.path.join(os.getcwd(), 'StatisResult', f'{src_dataset_name}', f'Result_{src_dataset_name}.xlsx'))

def excel_to_csv(path):
    result_excel = openpyxl.load_workbook(path)
    for sheet_name in result_excel.sheetnames:
        sheet = result_excel[sheet_name]
        csv_name = os.path.join(os.getcwd(), "StatisResult", "CAS", sheet_name + ".csv")
        with open(csv_name, "w", newline="") as csvfile:
            writer = csv.writer(csvfile)
            for row in sheet.iter_rows():
                writer.writerow([cell.value for cell in row])


if __name__ == "__main__":
    excel_to_csv('F://Code/ERPToolBox/StatisResult/CAS/Stastic_Result_CAS.xlsx')
    # transfer("CAS", "BA", 'C')
    # transfer("CAS", "TPR", 'D')
    # transfer("CAS", "FPR", 'E')
    # transfer("CAS", "F1", 'G')
    # transfer("CAS", "KAPPA", 'H')
    # transfer("CAS", "AUC", 'I')
